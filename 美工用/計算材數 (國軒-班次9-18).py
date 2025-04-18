#!/usr/bin/env python3
from tkinterdnd2 import TkinterDnD, DND_FILES
import tkinter as tk
from tkinter import messagebox
from PIL import Image
import subprocess
import os
import re
import xml.etree.ElementTree as ET
import sys
import fcntl
import json
from datetime import date, datetime
from openpyxl import load_workbook, Workbook
import socketio



# 在其他 import 下方加入全域變數定義
EMPLOYEE = "國軒"  # 請根據實際情況設定
SHARED_FOLDER = "/Volumes/助理美工區/#國軒"

# 將此處改成你自己的 ngrok 公開網址
NGROK_URL = 'https://onlycolor80193922.duckdns.org'

# 建立一個 Socket.IO client（開啟 debug 可以看更詳細 log）
sio = socketio.Client(logger=True, engineio_logger=True)

try:
    sio.connect(
        NGROK_URL,      # 例如 'https://773d-61-216-19-45.ngrok-free.app'
        # 不指定 transports，預設會自動 fallback 到 polling or websocket
        # 也不用指定 namespaces，預設就連根 namespace '/'
    )
except Exception as e:
    print("SocketIO 連線失敗:", e)
    sys.exit(1)

# 範例：複製 stats 到共享資料夾（如果有用到）
def copy_stats_to_shared(stats):
    if not os.path.exists(SHARED_FOLDER):
        print(f"共享資料夾 {SHARED_FOLDER} 不存在！")
        return
    shared_filename = os.path.join(SHARED_FOLDER, f"{EMPLOYEE}.json")
    try:
        with open(shared_filename, "w", encoding="utf8") as f:
            json.dump(stats, f, ensure_ascii=False, indent=4)
        print(f"成功將 stats 上傳到共享資料夾：{shared_filename}")
    except Exception as e:
        print("上傳 stats 到共享資料夾失敗:", e)

# 防止多重實例運行
def single_instance():
    try:
        lockfile = open("/tmp/my_app.lock", "w")
        fcntl.lockf(lockfile, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except IOError:
        print("應用程式已經在運行！")
        sys.exit()

single_instance()

# 取消 PIL 的最大像素限制
Image.MAX_IMAGE_PIXELS = None

####################################
# 讀取及儲存統計資料 (轉換為 Excel 檔案)
####################################
STATS_JSON_PATH = "/Users/onlycolor/Public/stats.json"

def load_stats(filename=STATS_JSON_PATH):
    if os.path.exists(filename):
        try:
            with open(filename, "r", encoding="utf8") as f:
                stats = json.load(f)
            # 修改統計資料結構：若 daily 中沒有班別資訊，則初始化為正班與加班兩個欄位
            if "daily" in stats:
                for day, data in stats["daily"].items():
                    if not isinstance(data, dict) or ("regular" not in data and "overtime" not in data):
                        stats["daily"][day] = {
                            "regular": {"file_count": data.get("file_count", 0), "material": data.get("material", 0.0)},
                            "overtime": {"file_count": 0, "material": 0.0}
                        }
            else:
                stats["daily"] = {}
            if "monthly" not in stats:
                stats["monthly"] = {}
            if "files" not in stats:
                stats["files"] = {}
            return stats
        except Exception as e:
            print("讀取統計檔案失敗:", e)
            return {"daily": {}, "monthly": {}, "files": {}}
    else:
        return {"daily": {}, "monthly": {}, "files": {}}

def save_stats(stats, filename=STATS_JSON_PATH):
    try:
        with open(filename, "w", encoding="utf8") as f:
            json.dump(stats, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print("儲存統計檔案失敗:", e)

def save_stats_to_excel(stats, filename="stats.xlsx", password="secret"):
    wb = Workbook()

    # 建立 Daily 工作表
    ws_daily = wb.active
    ws_daily.title = "Daily"
    ws_daily.append(["日期", "檔案數量", "總材數"])
    for day, data in stats.get("daily", {}).items():
        # 這裡可以根據需求決定如何呈現正班與加班的統計數據
        file_count = data.get("regular", {}).get("file_count", 0) + data.get("overtime", {}).get("file_count", 0)
        material = data.get("regular", {}).get("material", 0.0) + data.get("overtime", {}).get("material", 0.0)
        ws_daily.append([day, file_count, material])
    ws_daily.protection.sheet = True
    ws_daily.protection.password = password

    # 建立 Monthly 工作表
    ws_monthly = wb.create_sheet(title="Monthly")
    ws_monthly.append(["月份", "班別", "檔案數量", "總材數"])
    for month, data in stats.get("monthly", {}).items():
        # 如果沒有分班別結構（以防萬一），則以原有數值顯示
        if isinstance(data, dict) and "regular" in data and "overtime" in data:
            ws_monthly.append(
                [month, "正班", data["regular"].get("file_count", 0), data["regular"].get("material", 0.0)])
            ws_monthly.append(
                [month, "加班", data["overtime"].get("file_count", 0), data["overtime"].get("material", 0.0)])
        else:
            ws_monthly.append([month, "", data.get("file_count", 0), data.get("material", 0.0)])
    ws_monthly.protection.sheet = True
    ws_monthly.protection.password = password

    # 建立 Files 工作表
    ws_files = wb.create_sheet(title="Files")
    ws_files.append(["記錄鍵", "日期", "月份", "材數"])
    for key, data in stats.get("files", {}).items():
        ws_files.append([key, data.get("day", ""), data.get("month", ""), data.get("material", 0.0)])
    ws_files.protection.sheet = True
    ws_files.protection.password = password

    try:
        wb.save(filename)
        print(f"統計資料已儲存至受保護的 Excel 檔案：{filename}")
    except Exception as e:
        print("儲存統計 Excel 檔案失敗:", e)

####################################
# 解析 Excel 中的數量數值
# 優先讀取 V 欄 (第22欄, 0-index 21)
# 使用正則 search 抽取符合模式的數字—底各—數字（允許中間有空格）；
# 若 V 欄無效，則讀取 U 欄 (第21欄, 0-index 20) 為純數字（此時 allowed 固定為1）。
# 返回 (allowed, copies, source) 其中 source 為 "V" 或 "U"
####################################
def parse_quantity(row):
    pattern = re.compile(r'(\d+)\s*底各\s*(\d+)', re.UNICODE)
    v_val = row[21] if len(row) > 21 else None
    if v_val is not None:
        v_str = str(v_val).strip()
        match = pattern.search(v_str)
        if match:
            try:
                allowed = int(match.group(1))
                copies = int(match.group(2))
                return allowed, copies, "V"
            except Exception as e:
                print(f"正則解析 V 欄值 '{v_str}' 失敗: {e}")
        else:
            print(f"V 欄值 '{v_str}' 未抽取出符合格式的數字")
    u_val = row[20] if len(row) > 20 else None
    if u_val is not None:
        try:
            u_str = str(u_val).strip()
            if not re.fullmatch(r'\d+', u_str):
                raise ValueError("非純數字")
            copies = int(u_str)
            return 1, copies, "U"
        except Exception as e:
            print(f"解析 U 欄值 '{u_val}' 失敗: {e}")
    return 1, 1, None

####################################
# Excel 規則處理：讀取銷貨單號、順序、寬、高及數量資訊
####################################
def process_excel_file(excel_file):
    try:
        wb = load_workbook(excel_file)
        sheet = wb.active
        rules = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[20] is not None and isinstance(row[20], str) and row[20].strip() == "數量":
                continue
            if not row[0] or not row[1]:
                continue
            sales_order = str(row[0]).strip()
            sequence = str(row[1]).strip()
            try:
                width = float(row[18]) if row[18] is not None else None
                height = float(row[19]) if row[19] is not None else None
            except (ValueError, IndexError):
                width = height = None
            allowed, copies, source = parse_quantity(row)
            if sales_order and sequence and width is not None and height is not None and allowed is not None and copies is not None:
                if sales_order not in rules:
                    rules[sales_order] = {}
                rules[sales_order][sequence] = {
                    'width': width,
                    'height': height,
                    'allowed': allowed,
                    'copies': copies,
                    'source': source
                }
        return rules
    except Exception as e:
        print(f"解析 Excel 檔案時錯誤 {excel_file}: {e}")
        return {}

####################################
# 檔案名稱解析：取得 Excel 鍵與記錄鍵
####################################
def parse_file_keys(file_name):
    """
    從檔案名稱中解析出兩個鍵：
      - Excel 鍵：僅取檔案名稱（去除副檔名）前兩部分，例如 "23091119-001"
      - 記錄鍵：使用完整的檔案名稱（去除副檔名），例如 "23091119-001" 或 "23091119-001-@10" 或 "23091119-001-1"
    """
    base = file_name.rsplit('.', 1)[0]
    parts = base.split('-')
    if len(parts) < 2:
        return None, None
    excel_key = f"{parts[0]}-{parts[1]}"
    record_key = base
    return excel_key, record_key

####################################
# 取得檔案序號：若記錄鍵含第三部分則返回，否則預設為 1
####################################
def get_file_index(record_key):
    parts = record_key.split('-')
    if len(parts) >= 3:
        try:
            return int(parts[2])
        except:
            return 1
    return 1

####################################
# 單位轉換及尺寸讀取函數
####################################
def convert_px_to_cm(px_value):
    return px_value / 96 * 2.54

def get_ai_dimensions(ai_file):
    svg_file = os.path.splitext(ai_file)[0] + ".svg"
    try:
        subprocess.run([
            "inkscape",
            ai_file,
            "--export-filename=" + svg_file,
            "--export-area-page"
        ], check=True)
        tree = ET.parse(svg_file)
        root = tree.getroot()
        width_str = root.get("width")
        height_str = root.get("height")
        if width_str is None or height_str is None:
            viewBox = root.get("viewBox")
            if viewBox:
                values = viewBox.split()
                if len(values) == 4:
                    width = convert_px_to_cm(float(values[2]))
                    height = convert_px_to_cm(float(values[3]))
                else:
                    width = height = 0
            else:
                width = height = 0
        else:
            if "px" in width_str:
                width = convert_px_to_cm(float(width_str.replace("px", "").strip()))
            elif "cm" in width_str:
                width = float(width_str.replace("cm", "").strip())
            else:
                width = convert_px_to_cm(float(width_str.strip()))
            if "px" in height_str:
                height = convert_px_to_cm(float(height_str.replace("px", "").strip()))
            elif "cm" in height_str:
                height = float(height_str.replace("cm", "").strip())
            else:
                height = convert_px_to_cm(float(height_str.strip()))
        return width, height
    except Exception as e:
        print(f"解析 AI 檔案時錯誤 {ai_file}: {e}")
        return 0, 0
    finally:
        if os.path.exists(svg_file):
            try:
                os.remove(svg_file)
                print(f"已刪除臨時 SVG 檔案: {svg_file}")
            except Exception as e:
                print(f"刪除 SVG 檔案時錯誤: {e}")

def get_image_dimensions(file_path):
    try:
        with Image.open(file_path) as img:
            width_px, height_px = img.size
            dpi = img.info.get('dpi', (72, 72))
            width_cm = width_px / dpi[0] * 2.54
            height_cm = height_px / dpi[1] * 2.54
            return width_cm, height_cm
    except Exception as e:
        print(f"讀取圖片尺寸時錯誤 {file_path}: {e}")
        return 0, 0

####################################
# 拖放事件處理
####################################
def on_drop(event):
    files = [f.strip("{}") for f in event.data.split()]

    # ─── 新增：強制要帶 Excel ───
    excel_files = [f for f in files if f.lower().endswith('.xlsx')]
    if not excel_files:
        messagebox.showwarning("缺少 Excel 檔案", "請同時拖入至少一個 .xlsx 檔案，否則無法執行。")
        return
    # ────────────────────────────────
    stats = load_stats()
    non_xlsx_count = 0
    result_info = []
    allowed_extensions = ['.ai', '.jpg', '.jpeg', '.tiff', '.tif']
    today_str = date.today().strftime("%Y-%m-%d")
    month_str = date.today().strftime("%Y-%m")
    current_hour = datetime.now().hour

    # 判斷班別：假設每日 09:00～18:00 為正班，其餘為加班
    if 9 <= current_hour < 18:
        shift = "regular"
    else:
        shift = "overtime"
    print(f"目前班別：{shift} (目前小時: {current_hour})")

    # 檢查是否有 Excel 檔（用於規則比對）
    excel_file = None
    for file in files:
        if file.lower().endswith('.xlsx'):
            excel_file = file.strip("{}")
            break

    if excel_file:
        excel_base = os.path.basename(excel_file).rsplit('.', 1)[0]
        excel_sales_order = excel_base[:8]
        print(f"Excel 銷貨單號標準：{excel_sales_order}")
        rules = process_excel_file(excel_file)
        print(f"已從 Excel 取得規則：{rules}")
    else:
        print("未提供 Excel 檔案，將不進行尺寸比對！")
        rules = {}
        excel_sales_order = None

    # 初始化今日與本月統計（以班別區分），同時建立平面欄位以供網頁使用
    if today_str not in stats["daily"]:
        stats["daily"][today_str] = {
            "regular": {"file_count": 0, "material": 0.0},
            "overtime": {"file_count": 0, "material": 0.0},
            "file_count": 0, "material": 0.0
        }
    if month_str not in stats["monthly"]:
        stats["monthly"][month_str] = {
            "regular": {"file_count": 0, "material": 0.0},
            "overtime": {"file_count": 0, "material": 0.0},
            "file_count": 0, "material": 0.0
        }
    else:
        if not (isinstance(stats["monthly"][month_str], dict) and "regular" in stats["monthly"][month_str]):
            old_file_count = stats["monthly"][month_str].get("file_count", 0)
            old_material = stats["monthly"][month_str].get("material", 0.0)
            stats["monthly"][month_str] = {
                "regular": {"file_count": old_file_count, "material": old_material},
                "overtime": {"file_count": 0, "material": 0.0},
                "file_count": old_file_count, "material": old_material
            }

    TOLERANCE = 2.1  # 容差 2.1 公分

    for file in files:
        file = file.strip("{}")
        file_name = os.path.basename(file)
        lower_file = file.lower()
        if lower_file.endswith('.xlsx'):
            continue  # 不計算 Excel 資料檔
        if not any(lower_file.endswith(ext) for ext in allowed_extensions):
            print(f"不支援的檔案格式: {file}")
            continue

        # 取得 Excel 鍵與記錄鍵
        excel_key, record_key = parse_file_keys(file_name)
        if excel_key is None or record_key is None:
            print(f"無法解析檔名 {file_name}，跳過。")
            continue

        file_sales_order = excel_key.split('-')[0][:8]
        if excel_sales_order and file_sales_order != excel_sales_order:
            messagebox.showwarning("銷貨單號不匹配",
                f"檔案 {file_name} 的銷貨單號 {file_sales_order} 與 Excel 銷貨單號 {excel_sales_order} 不匹配，檔案被擋下。")
            print(f"檔案 {file_name} 銷貨單號不匹配，跳過。")
            continue

        # 取得檔案序號
        file_parts = record_key.split('-')
        if len(file_parts) >= 3:
            try:
                file_index = int(file_parts[2])
            except:
                file_index = 1
        else:
            file_index = 1

        # 檢查 Excel 規則（略過部份規則判斷）
        if rules:
            sales_order = excel_key.split('-')[0]
            seq = excel_key.split('-')[1]
            if sales_order not in rules or seq not in rules[sales_order]:
                messagebox.showwarning("資料不符", f"{file_name} 沒有對應的 Excel 規則，檔案被擋下。")
                print(f"檔案 {file_name} 的銷貨順序不符，跳過。")
                continue
            rule = rules[sales_order][seq]
            rule_width = rule.get("width")
            rule_height = rule.get("height")
            allowed_num = rule.get("allowed")
            copies = rule.get("copies")
            source = rule.get("source")
            if source == "U":
                rec_lower = record_key.lower()
                expected1 = excel_key.lower()
                expected2 = (excel_key + "-@10").lower()
                if rec_lower != expected1 and rec_lower != expected2:
                    messagebox.showwarning("檔案格式錯誤", f"{record_key} 此項次只有1底")
                    print(f"{record_key} 格式錯誤（來源 U），跳過。")
                    continue
            else:
                if allowed_num > 1:
                    if len(file_parts) < 3 or not re.fullmatch(r'\d+', file_parts[2]):
                        messagebox.showwarning("檔案格式錯誤", f"{record_key} 應包含有效的底次序號")
                        print(f"{record_key} 格式錯誤（需第三部分為純數字），跳過。")
                        continue
            if file_index > allowed_num:
                messagebox.showwarning("檔案數量錯誤", f"{record_key} 只有 {allowed_num} 底")
                print(f"{record_key} 序號超出允許數 ({allowed_num})，跳過。")
                continue
        else:
            rule_width = rule_height = None
            allowed_num = 1
            copies = 1

        # 取得檔案尺寸
        if lower_file.endswith('.ai'):
            width, height = get_ai_dimensions(file)
        elif lower_file.endswith(('.jpg', '.jpeg', '.tiff', '.tif')):
            width, height = get_image_dimensions(file)
        else:
            width, height = 0, 0

        multiplier = 10 if "@10" in lower_file else 1
        if multiplier == 10:
            effective_width = width * 10
            effective_height = height * 10
        else:
            effective_width = width
            effective_height = height

        if rule_width is not None and rule_height is not None:
            if not (abs(effective_width - rule_width) <= TOLERANCE and abs(effective_height - rule_height) <= TOLERANCE):
                messagebox.showwarning("尺寸比對錯誤",
                    f"{excel_key} 比對尺寸錯誤\n實際尺寸：{effective_width:.2f}×{effective_height:.2f} cm\n規定尺寸：{rule_width:.2f}×{rule_height:.2f} cm")
                print(f"檔案 {file_name} 尺寸不符 (實際: {effective_width:.2f}×{effective_height:.2f} cm, 規定: {rule_width:.2f}×{rule_height:.2f} cm)，請確認。")
                continue

        if multiplier == 10:
            computed_material = (width * height * 100) / 900
        else:
            computed_material = (width * height) / 900
        final_material = computed_material * (copies if copies is not None else 1)

        non_xlsx_count += 1
        result_info.append((record_key, effective_width, effective_height, final_material, copies))
        print(f"檔案: {record_key}  有效寬: {effective_width:.2f} cm, 有效高: {effective_height:.2f} cm, 材數: {final_material:.2f} (乘以 {copies} 倍)")

        if record_key in stats["files"]:
            messagebox.showinfo("檔案覆蓋", f"已將記錄中的 {record_key} 刪除 (數量 {copies} 個)")
            old_record = stats["files"][record_key]
            old_day = old_record.get("day")
            old_month = old_record.get("month")
            old_material = old_record.get("material", 0.0)
            if old_day in stats["daily"]:
                old_shift = old_record.get("shift", "regular")
                stats["daily"][old_day][old_shift]["file_count"] -= 1
                stats["daily"][old_day][old_shift]["material"] -= old_material
            if old_month in stats["monthly"]:
                old_shift = old_record.get("shift", "regular")
                if old_shift not in stats["monthly"][old_month]:
                    stats["monthly"][old_month][old_shift] = {"file_count": 0, "material": 0.0}
                stats["monthly"][old_month][old_shift]["file_count"] -= 1
                stats["monthly"][old_month][old_shift]["material"] -= old_material
            print(f"記錄 {record_key} 已更新，已扣除舊記錄。")

        stats["files"][record_key] = {"day": today_str, "month": month_str, "material": final_material, "shift": shift}
        stats["daily"][today_str][shift]["file_count"] += 1
        stats["daily"][today_str][shift]["material"] += final_material
        stats["monthly"][month_str][shift]["file_count"] += 1
        stats["monthly"][month_str][shift]["material"] += final_material

    # ★ 更新平面統計欄位：累加今日及本月各班數據
    daily_data = stats["daily"].get(today_str, {"regular": {"file_count": 0, "material": 0.0}, "overtime": {"file_count": 0, "material": 0.0}})
    total_daily_files = daily_data["regular"]["file_count"] + daily_data["overtime"]["file_count"]
    total_daily_material = daily_data["regular"]["material"] + daily_data["overtime"]["material"]
    stats["daily"][today_str]["file_count"] = total_daily_files
    stats["daily"][today_str]["material"] = total_daily_material

    monthly_data = stats["monthly"].get(month_str, {"regular": {"file_count": 0, "material": 0.0}, "overtime": {"file_count": 0, "material": 0.0}})
    total_monthly_files = monthly_data["regular"]["file_count"] + monthly_data["overtime"]["file_count"]
    total_monthly_material = monthly_data["regular"]["material"] + monthly_data["overtime"]["material"]
    stats["monthly"][month_str]["file_count"] = total_monthly_files
    stats["monthly"][month_str]["material"] = total_monthly_material

    # 儲存並更新 Excel 檔案、複製到共享資料夾
    save_stats(stats)
    save_stats_to_excel(stats, filename="/Users/onlycolor/Desktop/轉檔excel表/stats.xlsx", password="secret")
    copy_stats_to_shared(stats)

    # 組合結果訊息 (分正班/加班)
    # 今日統計
    today_regular = stats["daily"][today_str]["regular"]
    today_overtime = stats["daily"][today_str]["overtime"]
    # 本月統計
    month_regular = stats["monthly"][month_str]["regular"]
    month_overtime = stats["monthly"][month_str]["overtime"]

    msg = f"製作檔案數量: {non_xlsx_count}\n"
    for key, eff_w, eff_h, m, mult in result_info:
        msg += f"{key} : 寬 = {eff_w:.2f} cm, 高 = {eff_h:.2f} cm, 材數 = {m:.2f} (數量 {mult} 個)\n"

    msg += "\n[今日班別統計]\n"
    msg += f"正班: {today_regular['file_count']} 個檔案, 材數 = {today_regular['material']:.2f}\n"
    msg += f"加班: {today_overtime['file_count']} 個檔案, 材數 = {today_overtime['material']:.2f}\n"

    msg += "\n[本月班別統計]\n"
    msg += f"正班: {month_regular['file_count']} 個檔案, 材數 = {month_regular['material']:.2f}\n"
    msg += f"加班: {month_overtime['file_count']} 個檔案, 材數 = {month_overtime['material']:.2f}"

    messagebox.showinfo("計算結果", msg)

    # ★ 通知伺服器更新，讓連線的網頁自動刷新
    notify_update()

# 新增：利用 socketio 客戶端通知更新的函式
def notify_update():
    # 直接 emit 到 '/' 預設 namespace
    sio.emit('stats_update', {'message': 'Stats updated'})
    print("已發送更新通知")


####################################
# 將統計資料存入受保護的 Excel 檔案（保持原有排版）
####################################
def save_stats_to_excel(stats, filename="stats.xlsx", password="secret"):
    from openpyxl import Workbook
    wb = Workbook()

    # 建立 Daily 工作表
    ws_daily = wb.active
    ws_daily.title = "Daily"
    ws_daily.append(["日期", "檔案數量", "總材數"])
    for day, data in stats.get("daily", {}).items():
        file_count = data.get("regular", {}).get("file_count", 0) + data.get("overtime", {}).get("file_count", 0)
        material = data.get("regular", {}).get("material", 0.0) + data.get("overtime", {}).get("material", 0.0)
        ws_daily.append([day, file_count, material])
    ws_daily.protection.sheet = True
    ws_daily.protection.password = password

    # 建立 Monthly 工作表
    ws_monthly = wb.create_sheet(title="Monthly")
    ws_monthly.append(["月份", "檔案數量", "總材數"])
    for month, data in stats.get("monthly", {}).items():
        ws_monthly.append([month, data.get("file_count", 0), data.get("material", 0.0)])
    ws_monthly.protection.sheet = True
    ws_monthly.protection.password = password

    # 建立 Files 工作表
    ws_files = wb.create_sheet(title="Files")
    ws_files.append(["記錄鍵", "日期", "月份", "材數"])
    for key, data in stats.get("files", {}).items():
        ws_files.append([key, data.get("day", ""), data.get("month", ""), data.get("material", 0.0)])
    ws_files.protection.sheet = True
    ws_files.protection.password = password

    try:
        wb.save(filename)
        print(f"統計資料已儲存至受保護的 Excel 檔案：{filename}")
    except Exception as e:
        print("儲存統計 Excel 檔案失敗:", e)

####################################
# 建立視窗及拖放設定
####################################
from tkinterdnd2 import TkinterDnD
root = TkinterDnD.Tk()
root.configure(bg="white")
root.title("材數計算")
window_width = 600
window_height = 400
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
center_x = int((screen_width - window_width) / 2)
center_y = int((screen_height - window_height) / 2) - 100
root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
root.config(bg="#f5f5f5")

title_label = tk.Label(root, text="丟入轉檔檔案和資料 xlsx ai tiff jpg", bg="#f5f5f5", fg="black",
                       font=("Arial", 18, "bold"))
title_label.place(relx=0.5, y=80, anchor="center")

drop_area = tk.Label(root, text="丟入檔案", bg="white", fg="black", font=("Arial", 16, "bold"))
drop_area.place(x=100, y=140, width=400, height=200)
drop_area.drop_target_register(DND_FILES)
drop_area.dnd_bind('<<Drop>>', on_drop)

root.mainloop()