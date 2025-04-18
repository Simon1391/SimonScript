from tkinterdnd2 import TkinterDnD, DND_FILES
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import openpyxl
from PIL import Image
import sys
import fcntl

# 防止多重實例運行
def single_instance():
    try:
        # 嘗試創建一個鎖定文件
        lockfile = open("/tmp/my_app.lock", "w")
        fcntl.lockf(lockfile, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except IOError:
        print("應用程式已經在運行！")
        sys.exit()

single_instance()

## 取消 PIL 的最大像素限制
Image.MAX_IMAGE_PIXELS = None  # 取消限制

# 調整比例，無論檔名是否包含 @10 都乘以 10
def check_and_adjust_size(file_width, file_height, rule_width, rule_height):

    print(f"檢查尺寸：檔案寬度: {file_width}, 高度: {file_height}，規則寬度: {rule_width}, 規則高度: {rule_height}")

    # 檢查尺寸是否符合規則
    if file_width != rule_width or file_height != rule_height:
        print(f"尺寸不符，檔案寬度: {file_width}, 高度: {file_height}，規則寬度: {rule_width}, 規則高度: {rule_height}")

        # 如果尺寸不符，將文件尺寸調整 10 倍
        adjusted_width = file_width * 10
        adjusted_height = file_height * 10
        print(f"調整後的檔案寬度: {adjusted_width}, 調整後的檔案高度: {adjusted_height}")

        return adjusted_width, adjusted_height
    else:
        print("尺寸符合規則，不需要調整。")
        return file_width, file_height

# 比對尺寸的函數
def compare_dimensions(ai_width, ai_height, order_width, order_height, tolerance=0.2):
    # 计算误差
    width_diff = abs(ai_width - order_width)
    height_diff = abs(ai_height - order_height)

    if width_diff > tolerance or height_diff > tolerance:
        return False
    return True

# 处理 Excel 文件，返回规则字典
def process_excel_file(excel_file):
    wb = load_workbook(excel_file)
    sheet = wb.active  # 获取活动的 sheet

    rules = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过第一行表头
        if not row[0] or not row[1] or not row[2]:
            continue

        sales_order = str(row[0]).strip()
        sequence = str(row[1]).strip()
        product_spec = str(row[2]).strip()

        try:
            # 获取宽度和高度，确保它们存在并且有效
            width = float(row[18]) if row[18] and isinstance(row[18], (int, float)) else None  # 寬度在第19行
            height = float(row[19]) if row[19] and isinstance(row[19], (int, float)) else None  # 高度在第20行
        except (ValueError, IndexError):
            width = height = None

        board_material = str(row[10]).strip()  # 第11列是板材

        # 仅当銷貨單號、順序號、寬度和高度都有效时，才将规则添加到字典中
        if sales_order and sequence and product_spec and height is not None and width is not None:
            if sales_order not in rules:
                rules[sales_order] = {}

            if '板' in board_material:
                if (height > 300 and width <= 120) or (width > 300 and height <= 120):
                    product_spec = f"{product_spec}細長類搭接"
                elif height > 153 and width > 153:
                    product_spec = f"{product_spec}搭接"
            elif 'PVC' in product_spec and height > 153 and width > 153:
                product_spec = f"{product_spec}搭接"

            rules[sales_order][sequence] = {'product_spec': product_spec, 'height': height, 'width': width}

    return rules


# 定義常數區域
SCALE_FACTOR = 10  # 調整檔案尺寸的縮放因子，用於調整大小以符合規則，通常用於解決尺寸與規格之間的比對問題。

def adjust_for_scale(width, height, scale_factor):
    scale_factor = 10
    """
    根據縮放因子調整尺寸。

    :param width: 原始寬度
    :param height: 原始高度
    :param scale_factor: 縮放因子
    :return: 調整後的寬度和高度
    """
    new_width = width * scale_factor
    new_height = height * scale_factor

    # 顯示調整後的尺寸
    print(f"調整後的寬度: {new_width}, 調整後的高度: {new_height}")

    return new_width, new_height


def get_file_size(ai_file):
    # 轉換 AI 為 SVG
    svg_file = os.path.splitext(ai_file)[0] + ".svg"

    try:
        # 執行 Inkscape 轉換
        subprocess.run([
            "inkscape", ai_file, "--export-filename", svg_file
        ], check=True)

        # 解析 SVG 獲取尺寸
        tree = ET.parse(svg_file)
        root = tree.getroot()

        # 嘗試從 width / height 讀取尺寸
        width = root.get("width")
        height = root.get("height")

        if width and height:
            width = float(width.replace("cm", "").strip())
            height = float(height.replace("cm", "").strip())
        else:
            # 如果 width / height 無效，改用 viewBox
            viewBox = root.get("viewBox")
            if viewBox:
                values = viewBox.split()
                if len(values) == 4:
                    width = float(values[2]) / 100  # 轉為 cm
                    height = float(values[3]) / 100
                else:
                    raise ValueError("SVG viewBox 格式錯誤")

        return width, height

    except Exception as e:
        print(f"提取 AI 檔案尺寸時出錯: {e}")
        return None, None

    finally:
        # 確保刪除臨時 SVG 檔案
        if os.path.exists(svg_file):
            try:
                os.remove(svg_file)
                print(f"已刪除臨時 SVG 檔案: {svg_file}")
            except Exception as delete_error:
                print(f"無法刪除 SVG 檔案: {delete_error}")


def check_filename_for_at10(file_name):
    """檢查檔案名稱是否包含 @10"""
    if '@10' in file_name or '＠10' in file_name:
        return True
    return False


# 假設 inkscape 導出的尺寸是像素（px）
def convert_px_to_cm(px_value):
    # 假設屏幕分辨率是 96 DPI，1 英寸 = 96 像素
    return px_value / 96 * 2.54  # 轉換為厘米

# 处理 SVG 文件，提取尺寸并转换为厘米
def get_svg_dimensions(svg_file):
    tree = ET.parse(svg_file)
    root = tree.getroot()
    width = float(root.attrib['width'])
    height = float(root.attrib['height'])

    # 转换为厘米
    width_cm = convert_px_to_cm(width)
    height_cm = convert_px_to_cm(height)

    return width_cm, height_cm

import os
import shutil
import subprocess
import tkinter as tk
from tkinter import messagebox


def check_size_match(actual_width, actual_height, rule_width, rule_height, tolerance=0.2):
    width_diff = abs(actual_width - rule_width)
    height_diff = abs(actual_height - rule_height)

    # 如果寬度和高度的誤差都在容忍範圍內，則視為匹配
    if width_diff <= tolerance and height_diff <= tolerance:
        print(f"尺寸匹配，檔案符合規則：寬度 = {actual_width}, 高度 = {actual_height}")
        return True
    else:
        print(f"尺寸不符，誤差範圍：寬度誤差 = {width_diff}, 高度誤差 = {height_diff}")
        return False

# 假設 Excel 解析後得到的品名規格與資料夾名稱對應
product_specs = {
    ('25031813', '001'): 'Folder_A',
    ('25031813', '002'): 'Folder_B',
}

def get_product_spec_from_excel(excel_file):
    """
    這個函數會從提供的 Excel 檔案中解析出銷貨單號與順序號對應的品名規格。
    返回的是一個字典，鍵是銷貨單號與順序號的組合，值是對應的品名規格。

    參數:
        excel_file: Excel 檔案路徑，用來提取產品規格信息

    返回:
        dict: 包含銷貨單號與順序號對應的品名規格字典
    """
    # 在這裡解析 Excel 檔案，並生成對應的字典
    # 假設以下是生成字典的步驟
    product_specs = {}  # 一個空字典來存放銷貨單號與品名規格的對應關係

    # 這裡會加上讀取 excel 的邏輯，例如用 pandas 來解析 Excel
    # 假設 Excel 中有列出銷貨單號與順序號，我們從這些數據來組合字典

    # 返回解析出來的字典
    return product_specs

def move_file_to_folder(file_path, folder_name):
    # 根據對應的資料夾名稱移動檔案
    target_folder = f"/Users/onlycolor/Desktop/{folder_name}"
    os.makedirs(target_folder, exist_ok=True)  # 確保資料夾存在
    target_path = os.path.join(target_folder, os.path.basename(file_path))

    # 嘗試移動檔案，並確保即使發生錯誤也會刪除 SVG 檔案
    try:
        shutil.move(file_path, target_path)
        print(f'檔案已移動到 {target_path}')
    except Exception as e:
        print(f"移動檔案時出錯: {e}")
    finally:
        # 確保刪除臨時 SVG 檔案
        svg_file = os.path.splitext(file_path)[0] + ".svg"
        if os.path.exists(svg_file):
            try:
                os.remove(svg_file)
                print(f"已刪除臨時 SVG 檔案: {svg_file}")
            except Exception as e:
                print(f"刪除 SVG 檔案時出錯: {e}")

def check_and_get_product_spec(filename, product_specs):
    sales_order, order_number = extract_sales_order_and_number(filename)
    if sales_order and order_number:
        product_spec = product_specs.get((sales_order, order_number))
        if product_spec:
            print(f'檔案 {filename} 對應的品名規格是: {product_spec}')
            return product_spec
        else:
            print(f'無法找到 {sales_order}-{order_number} 的品名規格。')
            return None
    else:
        print(f'檔案 {filename} 的銷貨單號和順序號格式錯誤。')
        return None

def on_drop(event):
    files = event.data.split()  # 分割檔案路徑

    # 找到 Excel 檔案
    excel_file = None
    for file in files:
        if file.endswith('.xlsx'):
            excel_file = file
            break

    if excel_file:
        # 讀取 Excel 檔案並提取品名規格
        print(f'處理 Excel 檔案: {excel_file}')
        product_specs = get_product_spec_from_excel(excel_file)

        # 處理其他檔案
        for file in files:
            if file.endswith('.ai'):
                filename = os.path.basename(file)
                print(f'正在處理檔案: {filename}')
                product_spec = check_and_get_product_spec(filename, product_specs)
                if product_spec:
                    # 根據品名規格移動檔案
                    move_file_to_folder(file, product_spec)
                else:
                    print(f'錯誤: 無法處理檔案 {filename}，因為找不到對應的品名規格。')
            else:
                print(f'忽略檔案: {file}')


def get_product_spec(excel_file, sales_order, sequence_number):
    # 讀取 Excel 檔案
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    # 遍歷 Excel 中的每一行，假設銷貨單號在第 A 欄，順序號在 B 欄，品名規格在 C 欄
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
        order_number = row[0].value  # 銷貨單號
        seq_number = row[1].value  # 順序號
        product_spec = row[2].value  # 品名規格

        # 檢查是否匹配銷貨單號和順序號
        if str(order_number) == str(sales_order) and str(seq_number) == str(sequence_number):
            return product_spec  # 返回對應的品名規格

    return None  # 如果沒有匹配的品名規格，返回 None

# 使用 PIL 提取圖片物理尺寸 (單位: cm)
def get_physical_size(file_path):
    """取得圖片的物理尺寸 (單位：cm)"""
    img = Image.open(file_path)

    # 取得圖片的 DPI（每英寸多少像素）
    dpi = img.info.get('dpi', (72, 72))  # 如果沒有 DPI 信息，預設為 72
    width, height = img.size  # 獲取圖片的像素尺寸

    # 計算物理尺寸
    physical_width = width / dpi[0] * 2.54  # 將英寸轉換為厘米
    physical_height = height / dpi[1] * 2.54

    return physical_width, physical_height

def check_image_size(file_path, sales_order, sequence, rules):
    """比對圖片尺寸與規則是否符合"""
    # 取得圖片的物理尺寸
    width, height = get_physical_size(file_path)
    print(f"檢查圖片檔案: {file_path}，物理尺寸 - 寬度: {width} cm, 高度: {height} cm")

    # 根據銷貨單號和順序號從規則中取得對應的寬度和高度
    rule_width = rules.get(sales_order, {}).get(sequence, {}).get('width')
    rule_height = rules.get(sales_order, {}).get(sequence, {}).get('height')

    if rule_width and rule_height:
        tolerance = 0.2  # 容忍範圍

        # 比對尺寸
        if abs(width - rule_width) <= tolerance and abs(height - rule_height) <= tolerance:
            print(f"尺寸符合規則，檔案: {file_path}")
            return True
        else:
            print(f"尺寸不符，檔案: {file_path}")
            return False
    else:
        print(f"規則中找不到對應的尺寸，檔案: {file_path}")
        return False


def move_files_to_folder(files, rules, target_dir):
    inkscape_path = '/Applications/Inkscape.app/Contents/MacOS/inkscape'  # Inkscape 路徑

    # 用來儲存所有產生的 SVG 檔案
    svg_files = []

    for file in files:
        try:
            file_name = os.path.basename(file)
            file_parts = file_name.lower().split('-')

            if len(file_parts) < 2:
                print(f"警告: 檔名格式錯誤，無法解析: {file_name}")
                continue  # 跳過這個檔案

            sales_order = file_parts[0]  # 銷貨單號
            sequence = file_parts[1].replace(".ai", "").replace(".jpg", "").replace(".jpeg", "").replace(".tiff", "").replace(".tif", "")  # 順序號

            # 比對檔案名稱的銷貨單號和順序號
            if sales_order not in rules:
                print(f"警告: 檔案 {file_name} 的銷貨單號與資料不符")
                messagebox.showwarning("銷貨單號錯誤", f"{sales_order}-{sequence} 銷貨單號錯誤")
                continue  # 跳過這個檔案

            if sequence not in rules[sales_order]:
                print(f"警告: 檔案 {file_name} 的順序號與資料不符")
                messagebox.showwarning("順序號錯誤", f"{sales_order}-{sequence} 沒有第{sequence}項次")
                continue  # 跳過這個檔案

            print(f"檢查銷貨單號: {sales_order}，順序號: {sequence}")

            # 根據檔案格式選擇處理方式
            if file.lower().endswith('.ai'):
                # 處理 AI 檔案
                svg_file = file.replace(".ai", ".svg")
                subprocess.run([inkscape_path, file, "--export-plain-svg=" + svg_file])
                print(f"Inkscape 命令已執行: {inkscape_path} {file} --export-plain-svg={svg_file}")
                svg_files.append(svg_file)  # 記錄產生的 SVG 檔案

                ai_width, ai_height = get_svg_dimensions(svg_file)
                print(f"AI 檔案寬度: {ai_width}, 高度: {ai_height}")

                # 比對 AI 尺寸
                rule_width = rules[sales_order][sequence]['width']
                rule_height = rules[sales_order][sequence]['height']
                tolerance = 0.2  # 容忍範圍

                # 第一次比對
                if abs(ai_width - rule_width) <= tolerance and abs(ai_height - rule_height) <= tolerance:
                    print(f"尺寸符合規則，檔案: {file_name}")

                    # 檢查是否有 @10
                    if '@10' not in file_name:
                        # 根據 product_spec 創建資料夾
                        product_spec = rules[sales_order][sequence]['product_spec']
                        folder_path = os.path.join(target_dir, product_spec)

                        # 檢查資料夾是否已經存在，若已存在則不通知
                        if not os.path.exists(folder_path):
                            os.makedirs(folder_path)
                            print(f"資料夾 '{product_spec}' 已建立。")
                            messagebox.showinfo("資料夾建立", f"資料夾 '{product_spec}' 已建立。\n請通知美工增加新品項。")

                        target_file = os.path.join(folder_path, file_name)
                        shutil.copy2(file, target_file)  # 覆蓋檔案
                        print(f"已複製檔案: {file} 至 {target_file}")
                    else:
                        print(f"檔案 {file_name} 已經有 @10，跳過移動操作")
                        messagebox.showinfo("檔案不需要加 @10", f"檔案 {file_name} 不需要加 @10")
                else:
                    print(f"AI 檔案尺寸不符，檔案: {file_name}")
                    # **第二次比對**
                    ai_width *= 10
                    ai_height *= 10
                    print(f"進行第二次比對，AI 寬度: {ai_width}, 高度: {ai_height}")

                    if abs(ai_width - rule_width) <= tolerance and abs(ai_height - rule_height) <= tolerance:
                        print(f"第二次比對尺寸符合規則，檔案: {file_name}")
                        if '@10' in file_name:
                            # 根據 product_spec 創建資料夾
                            product_spec = rules[sales_order][sequence]['product_spec']
                            folder_path = os.path.join(target_dir, product_spec)

                            # 檢查資料夾是否已經存在，若已存在則不通知
                            if not os.path.exists(folder_path):
                                os.makedirs(folder_path)
                                print(f"資料夾 '{product_spec}' 已建立。")
                                messagebox.showinfo("資料夾建立", f"資料夾 '{product_spec}' 已建立。\n請通知美工增加新品項。")

                            target_file = os.path.join(folder_path, file_name)
                            shutil.copy2(file, target_file)  # 覆蓋檔案
                            print(f"已複製檔案: {file} 至 {target_file}")
                        else:
                            print(f"警告: 檔案 {file_name} 缺少 @10")
                            messagebox.showwarning("缺少 @10", f"檔案 {file_name} 沒有 @10，請檢查！")
                            continue  # 跳過此檔案，進行下一個檔案
                    else:
                        print(f"第二次比對尺寸仍不符，檔案: {file_name}")
                        messagebox.showwarning("檔案尺寸錯誤", f"檔案尺寸與規則不符: {file_name}")
                        continue  # 跳過此檔案，進行下一個檔案

            elif file.lower().endswith(('.jpg', '.jpeg', '.tiff', '.tif')):  # 支援 JPG, JPEG, TIFF, TIF
                # 處理圖片檔案
                width, height = get_physical_size(file)
                print(f"檢查圖片檔案: {file}，物理尺寸 - 寬度: {width} cm, 高度: {height} cm")

                # 根據銷貨單號和順序號從規則中取得對應的寬度和高度
                rule_width = rules.get(sales_order, {}).get(sequence, {}).get('width')
                rule_height = rules.get(sales_order, {}).get(sequence, {}).get('height')

                if rule_width and rule_height:
                    tolerance = 0.2  # 容忍範圍

                    # 第一次比對
                    if abs(width - rule_width) <= tolerance and abs(height - rule_height) <= tolerance:
                        print(f"尺寸符合規則，檔案: {file_name}")
                        # 檢查是否有 @10
                        if '@10' not in file_name:
                            # 根據 product_spec 創建資料夾
                            product_spec = rules[sales_order][sequence]['product_spec']
                            folder_path = os.path.join(target_dir, product_spec)

                            # 檢查資料夾是否已經存在，若已存在則不通知
                            if not os.path.exists(folder_path):
                                os.makedirs(folder_path)
                                print(f"資料夾 '{product_spec}' 已建立。")
                                messagebox.showinfo("資料夾建立", f"資料夾 '{product_spec}' 已建立。\n請通知美工增加新品項。")

                            target_file = os.path.join(folder_path, file_name)
                            shutil.copy2(file, target_file)  # 覆蓋檔案
                            print(f"已複製檔案: {file} 至 {target_file}")
                        else:
                            print(f"檔案 {file_name} 已經有 @10，跳過移動操作")
                            messagebox.showinfo("檔案不需要加 @10", f"檔案 {file_name} 不需要加 @10")
                    else:
                        print(f"圖片檔案尺寸不符，檔案: {file_name}")
                        # **第二次比對**
                        width *= 10
                        height *= 10
                        print(f"進行第二次比對，圖片寬度: {width}, 高度: {height}")

                        if abs(width - rule_width) <= tolerance and abs(height - rule_height) <= tolerance:
                            print(f"第二次比對尺寸符合規則，檔案: {file_name}")
                            if '@10' in file_name:
                                # 根據 product_spec 創建資料夾
                                product_spec = rules[sales_order][sequence]['product_spec']
                                folder_path = os.path.join(target_dir, product_spec)

                                # 檢查資料夾是否已經存在，若已存在則不通知
                                if not os.path.exists(folder_path):
                                    os.makedirs(folder_path)
                                    print(f"資料夾 '{product_spec}' 已建立。")
                                    messagebox.showinfo("資料夾建立", f"資料夾 '{product_spec}' 已建立。\n請通知美工增加新品項。")

                                target_file = os.path.join(folder_path, file_name)
                                shutil.copy2(file, target_file)  # 覆蓋檔案
                                print(f"已複製檔案: {file} 至 {target_file}")
                            else:
                                print(f"警告: 檔案 {file_name} 缺少 @10")
                                messagebox.showwarning("缺少 @10", f"檔案 {file_name} 沒有 @10，請檢查！")
                                continue  # 跳過此檔案，進行下一個檔案
                        else:
                            print(f"第二次比對尺寸仍不符，檔案: {file_name}")
                            messagebox.showwarning("檔案尺寸錯誤", f"圖片檔案 {file_name} 的尺寸與規則不符，請檢查！")
                else:
                    print(f"規則中找不到對應的尺寸，檔案: {file_name}")
                    messagebox.showwarning("檔案尺寸錯誤", f"規則中找不到對應的尺寸，檔案: {file_name}")
            else:
                print(f"不支持的檔案格式: {file_name}")

        except Exception as e:
            print(f"處理檔案 {file_name} 時發生錯誤: {e}")
            continue  # 若出現錯誤，跳過此檔案並繼續處理其他檔案

    # 刪除所有生成的 SVG 檔案
    for svg_file in svg_files:
        if os.path.exists(svg_file):
            os.remove(svg_file)
            print(f"已刪除 SVG 檔案: {svg_file}")


def on_drop(event):
    dropped_files = event.data.split()
    print(f"Dropped files: {dropped_files}")  # 打印拖放的文件列表

    ods_file = None
    excel_file = None
    files_to_move = []
    target_dir = None  # 確保 target_dir 初始化

    # 限制接受的檔案類型 (不包含 .psd)
    allowed_extensions = ['.xlsx', '.ai', '.tiff', '.jpg', '.jpeg', '.tif']
    invalid_files = []  # 儲存無效檔案

    for file in dropped_files:
        print(f"Processing file: {file}")  # 打印每个文件路径

        # 檢查文件是否符合允許的格式
        if any(file.lower().endswith(ext) for ext in allowed_extensions):
            if file.lower().endswith('.xlsx'):
                excel_file = file
            elif file.lower().endswith('.ai'):
                files_to_move.append(file)
            elif file.lower().endswith(('.tiff', '.jpg', '.jpeg', '.tif')):
                files_to_move.append(file)
        else:
            # 提取檔案名稱和副檔名
            filename = file.split('/')[-1]
            file_extension = filename.split('.')[-1].lower()
            # 加入無效檔案到列表，顯示具體的錯誤提示
            print(f"忽略檔案: {filename}，不符合支援的格式")
            invalid_files.append(f"{filename} ({file_extension}檔)")  # 記錄檔案名稱和副檔名

    # 如果有不符合格式的檔案，顯示錯誤通知
    if invalid_files:
        for invalid_file in invalid_files:
            messagebox.showwarning("禁止檔案格式", f"禁止丟入 {invalid_file}")
        return

    # 處理符合格式的檔案
    if ods_file:
        rules = process_ods_file(ods_file)
        target_dir = "/Users/onlycolor/Desktop/品名規格資料夾"
        move_files_to_folder(files_to_move, rules, target_dir)
    elif excel_file:
        rules = process_excel_file(excel_file)
        target_dir = "/Users/onlycolor/Desktop/品名規格資料夾"
        move_files_to_folder(files_to_move, rules, target_dir)
    else:
        print("未找到 ODS 或 Excel 檔案!")
        messagebox.showwarning("文件錯誤", "未找到 ODS 或 Excel 檔案")

    # 確保 target_dir 被正確設置
    if target_dir is None:
        print("錯誤: 目標資料夾未設定!")
        messagebox.showwarning("錯誤", "目標資料夾未設定，請檢查檔案類型")




# 建立視窗
root = TkinterDnD.Tk()
root.configure(bg="white")  # 設定視窗背景為白色
root.title("拖放 資料 檢查移動的檔案")

# 設定視窗尺寸
window_width = 600
window_height = 400

# 取得螢幕寬高
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# 計算置中再往上
center_x = int((screen_width - window_width) / 2)
center_y = int((screen_height - window_height) / 2) - 100  # 上移 100 像素

# 設定視窗位置與大小
root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

# 設定窗口背景顏色為淡灰色
root.config(bg="#f5f5f5")  # 設置窗口的背景顏色為淡灰色

# 設定標題文字，顯示檔案類型，並將字體增大
title_label = tk.Label(root, text="丟入轉檔檔案和資料 xlsx ai tiff jpg", bg="#f5f5f5", fg="black", font=("Arial", 18, "bold"))
title_label.place(relx=0.5, y=80, anchor="center")  # 使用 relx 置中顯示，並設定距離上方80像素

# 設定"丟入文件"框框，字體使用 Arial 並加粗
drop_area = tk.Label(root, text="丟入檔案", bg="white", fg="black", font=("Arial", 16, "bold"))
drop_area.place(x=100, y=140, width=400, height=200)  # 設定位置和大小，位置不變

# 綁定拖放事件
drop_area.drop_target_register(DND_FILES)
drop_area.dnd_bind('<<Drop>>', on_drop)

root.mainloop()
